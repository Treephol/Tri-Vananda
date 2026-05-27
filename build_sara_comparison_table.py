"""
SARA Brand Research — 30-brand × 9-column comparison table
Tier-S quality bar: world-class branding agency benchmark
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── PALETTE (SARA / Tri Vananda visual system) ─────────────────────────
INK         = "1A1A1A"   # primary text
FOREST      = "1F3A2E"   # SARA forest green (header)
GOLD        = "B89968"   # SARA gold accent
SAND        = "F5EFE5"   # category band
CREAM       = "FAF7F2"   # zebra row
WHITE       = "FFFFFF"
RULE        = "D6CFC2"   # rule line

# ─── FONT ───────────────────────────────────────────────────────────────
FONT_NAME = "Calibri"   # universal Excel-safe; falls back gracefully

# ─── DATA ───────────────────────────────────────────────────────────────
HEADERS = [
    "Brand",
    "Essence (5 words)",
    "Visual DNA",
    "Narrative Anchor",
    "Signature Ritual",
    "Audience",
    "Price / Access",
    "Defensibility Moat",
    "Lesson for SARA",
]

# Each row: (category, brand, essence, visual_dna, narrative, ritual, audience, price, moat, lesson)
# Categories grouped — category cell only present on first brand of group
DATA = [
    # ─── 1. VITALITY & LONGEVITY ────────────────────────────────────────
    ("Vitality & Longevity",
     "Clinique La Prairie",
     "Swiss longevity laboratory for self-extension.",
     "Clinical white + glacial blue. Helvetica-clean typography. Lake-and-Alps photography. Hexagonal cellular motif. White coats and marble. Scientific imagery dominant; hospitality imagery secondary.",
     "Science of longevity born in 1931 with Dr Niehans' revitalisation cell therapy — 90+ years of medical innovation made luxurious. Heritage is the product.",
     "The 7-day Revitalisation Programme — proprietary multi-disciplinary medical protocol combining cellular therapies, genetic diagnostics, and a 60+-specialist clinical team.",
     "Heads of state, royal families, global CEOs aged 50-75 seeking biological renewal. Reported 30% annual return rate.",
     "CHF 35,000-50,000 per week. Medical referral and consultation required. Multi-week stays standard.",
     "90+ years of clinical IP. FDA-impossible-elsewhere protocols. Doctor-to-guest ratio. Swiss medical legitimacy unrepeatable by pure hospitality brands.",
     "Medical legitimacy is the deepest moat in wellness. Partner authenticity (CLP) > performance claims. Build the science narrative through real protocols, not marketing language."),

    ("",
     "SHA Wellness Clinic",
     "Macrobiotic medicine for measurable transformation.",
     "White minimalist Mediterranean architecture. Geometric ziggurat silhouette. Soft sand + sage palette. Futurist-meets-monastic; fitness-clinical hybrid.",
     "East-meets-West integrative medicine. Combines traditional Eastern philosophy (macrobiotics, TCM) with Western diagnostic science for measurable outcomes.",
     "The SHA Method® — fully personalised diagnostic-driven multi-week programme integrating nutrition, mind, fitness, beauty, and healthy ageing.",
     "HNWIs 40-65, often founders or executives, often referred by previous guests. SHA Mexico (2024) now serving the Americas.",
     "€5,000-25,000 per week depending on programme tier. Clinical intake required. Multi-week stays preferred.",
     "Proprietary SHA Method® IP. Alfonso Bataller family ownership. Multi-property platform (Spain + Mexico) sharing one methodology.",
     "Name your method. A proprietary '[Brand] Method' turns an experience into a transferable asset. SARA needs a named protocol that can be referenced, taught, and exported."),

    ("",
     "Lanserhof",
     "F.X. Mayr precision for modern bodies.",
     "Architectural minimalism (Ingenhoven-designed Tegernsee and Sylt). Wood + glass + green. Monochrome restraint. Modernist medical-luxury hybrid; buildings function as brand collateral.",
     "The F.X. Mayr Cure modernised — diagnostic medicine of the gut as the foundation of all health, executed at the precision of a Swiss watch.",
     "Lanserhof Concept — diagnostic intake + Mayr fasting cure + manual abdominal treatment + chewing-discipline regime + alkaline-water rituals.",
     "German-speaking elite. Northern European executives. Often returning annually. Deeply discreet, low-publicity clientele.",
     "€4,500-12,000 per week. Minimum 7 nights, ideally 14-21. Medical screening required.",
     "Modernised Mayr IP. Multi-property European footprint (Tegernsee, Sylt, Lans, London). Award-winning architecture as silent brand asset.",
     "Architecture as brand. Lanserhof's buildings ARE the brand book. SARA's Phuket buildings must do narrative work — not just house programmes."),

    # ─── 2. ATHLETICS & PRACTICE ────────────────────────────────────────
    ("Athletics & Practice",
     "Equinox",
     "Performance fashion for the urban athlete.",
     "Black, white, and blood-red. Brutalist gym architecture. Provocative campaign photography (Steven Klein era). Athletic editorial. 'It's Not Fitness. It's Life.' typography.",
     "Fitness as identity, not chore — for people who don't just work out, they BUILD themselves. Provocation as differentiation.",
     "Tier X personal training, EQX Body methodology, branded recovery (CryoX, Pursuit cycling), and a bath-and-grooming amenity layer that elevates 'the locker room' to spa.",
     "28-50 urban professionals. $150K+ income. Design-conscious and status-aware. Aspirational mid-tier.",
     "$260-360/month gym tier. Equinox Hotel NYC $750+/night. Hudson Yards club ~$40K initiation.",
     "Real-estate-as-brand (90+ clubs in prestige locations). In-house creative team. Scale of media spend. Hotel extension proving the thesis.",
     "Provocation builds tribe. Equinox is divisive on purpose — and rich because of it. SARA should not try to please everyone; sharp positioning beats broad appeal."),

    ("",
     "Heimat",
     "European bathhouse culture, Los Angeles edition.",
     "Earth-tone monochrome (cream, terracotta, sand). Mediterranean modernism. Vintage athletic photography. Custom millwork. No visible logos. Art-curated walls.",
     "A members' club where movement, recovery, and food belong in one place — European Bäder culture imported to West Hollywood.",
     "The Heimat Bathhouse Circuit — sauna, steam, cold plunge, hammam, and infrared — integrated with training floor and Mediterranean café.",
     "LA creative-professional class, 30-50. Design-literate. Frustrated by Equinox's mass-market scale. Capped membership ~3,500.",
     "$4,500 initiation + $695/month. Active waitlist.",
     "Curated membership scarcity. Real-estate developer Sandow backing. Design execution (Walker Workshop) gym-chains can't replicate.",
     "Members' club logic > resort logic. Membership creates community, recurring revenue, and tribal behaviour. SARA Membership should be capped, curated, and conferring."),

    ("",
     "Tracy Anderson Method",
     "Cult method for the long, lean body.",
     "Cool grey + dancer-pink + brushed gold. Soft-lit studio photography. Gwyneth-adjacent celebrity association. Founder-as-face branding.",
     "One method, fifteen years of body research — a science of small-muscle, multi-vector movement that 'rebuilds the body from connective tissue up'.",
     "The Tracy Anderson Method — proprietary daily 60-minute choreographed sequence delivered live, on-demand, and at flagship studios.",
     "Affluent women 30-55. Fashion-adjacent. Results-obsessed. Often referred by friend or celebrity.",
     "$1,500/year online. $2,500+ for live classes. Flagship studio memberships $4,000+.",
     "15-year founder-led research. Gwyneth Paltrow partnership. Proprietary choreography IP. Single-method focus in a market of dilettantes.",
     "Founder credibility transfers. Tracy IS the brand. SARA needs identifiable founder voices (Khun Anudej, lead practitioners) so the brand has a face, not just a place."),

    # ─── 3. RECOVERY & BODYWORK ─────────────────────────────────────────
    ("Recovery & Bodywork",
     "Remedy Place",
     "Social wellness for the recovery generation.",
     "Apothecary cream + sage green + brass. Apple-store minimalism applied to ice baths. Clinical-warm hybrid. Branded ritual signage. Instagram-engineered visual design.",
     "Wellness as social currency — the first 'social wellness club' where biohacking becomes a shared luxury experience instead of a solo discipline.",
     "Group ice-bath sessions (3-minute contrast therapy), hyperbaric oxygen chambers, IV drips, cryotherapy — all bookable as social events with friends.",
     "25-45 high-achiever creatives, founders, athletes in NYC and LA. Pre-meeting / post-workout demographic.",
     "$375-700/month membership. $80-300 per modality drop-in. Founding member tier $1,500/month.",
     "First-mover in 'social wellness' category. Founder Dr Jonathan Leary as media figure. Hospitality-grade brand polish on biohacking modalities.",
     "Reframe the category. Remedy didn't invent ice baths — they invented the social context for ice baths. SARA can reframe wellness from 'treatment' to 'lifestyle ritual' similarly."),

    ("",
     "AIRE Ancient Baths",
     "Candlelit baths in ancient architecture.",
     "Stone, candlelight, water, shadow. Almost no logo, almost no copy. Sensory-led photography. Low-light always.",
     "Roman, Greek, and Ottoman bath traditions revived inside restored historical buildings (former mills, theatres, mansions) across global cities.",
     "The Thermal Circuit — Tepidarium, Caldarium, Frigidarium, Flotarium — plus optional wine bath, halotherapy salt bath, and candlelit massage.",
     "25-55 urban couples and friends seeking date-night sensual luxury. Not a clinical wellness audience.",
     "€60-280 per circuit + treatments. Bookable per visit. No membership required.",
     "Real-estate moat (each location is a heritage building you can't replicate). Patented Flotarium experience. Multi-city footprint (NYC, London, Madrid, Chicago, Copenhagen, Toronto).",
     "Atmosphere is the product. AIRE sells darkness, candlelight, and time — not 'treatments'. SARA's ritual spaces should be designed for sensory immersion first, programming second."),

    ("",
     "Bathhouse (Brooklyn)",
     "Concrete cathedral for cold-warm contrast.",
     "Raw concrete. Brutalist Brooklyn-warehouse architecture. Soft warm lighting. No-logo restraint. Monochrome graphic system. Helvetica discipline.",
     "A modern bathhouse for the city — a third place between gym and spa, designed like a nightclub but functioning like a sanctuary.",
     "Hot-cold contrast loop (sauna, steam, cold plunge, hammam) plus restaurant and cocktail bar. Full-day social use across one building.",
     "25-40 Brooklyn creatives in fashion and tech. Date-friendly, couple-friendly, group-friendly.",
     "$50-65 day pass. $200/month membership. ~3,000 visits per week per location.",
     "Architectural design (Verona Carpenter). No membership-only barrier — accessible volume model. F&B integration. Price accessibility at this aesthetic level.",
     "Hybrid use-cases multiply revenue. Bathhouse is gym + spa + restaurant + bar in one building. SARA properties should support breakfast → workout → bath → dinner in one continuous arc."),

    # ─── 4. MIND & COGNITION ────────────────────────────────────────────
    ("Mind & Cognition",
     "Open",
     "Movement and breath, designed beautifully.",
     "Warm peach + cream + clay. Soft-lit serif typography (custom). Generous whitespace. Hand-illustrated mark. Premium app UI craft.",
     "Meditation, breathwork, movement, and sound combined into one practice for the modern body that doesn't believe in 'just sitting still'.",
     "Open's signature class blends five modalities in 45 minutes — breath + movement + meditation + sound + stillness — delivered live, on-demand, and at LA flagship.",
     "25-45 design-literate professionals frustrated by Headspace and Calm's gamification and clinical aesthetic.",
     "$25/month app. $35-50 per studio class (LA flagship). $200 annual.",
     "Design quality. Founder Manoj Dias' credibility as a former meditation teacher. Brand restraint inside an oversaturated meditation market.",
     "Restraint as positioning. Open isn't trying to do everything — it just refuses to do meditation badly. SARA should under-deliver content noise to over-deliver atmosphere."),

    ("",
     "The School of Life",
     "Emotional intelligence as cultural project.",
     "Mid-century illustration. Mustard + ink + cream. Classic serif (Caslon-feel). Bookish editorial layout. Hand-drawn whimsy. Wes-Anderson-adjacent.",
     "Founded by Alain de Botton — philosophy made practical, emotional skills made teachable. 'Develop your emotional intelligence.'",
     "Classes (in-person + online), books, journals, conversation cards, therapy referrals — all designed as a curriculum for being human.",
     "30-55 educated professionals worldwide. Book-buying, gallery-going, therapy-positive demographic.",
     "£15-30 per class. £200+ courses. Books £10-25. Physical shops in London, Amsterdam, Berlin, Seoul.",
     "Alain de Botton author authority. 15+ years of content IP. Illustration-style consistency. Books-as-products revenue multiplier.",
     "Content is the long-tail moat. Tracy Anderson Method has classes. School of Life has books. SARA needs a publishable content layer — books, films, podcasts — that lives outside the property."),

    ("",
     "Hoffman Institute",
     "Eight-day deep emotional reset.",
     "Calm sage green + warm cream. Photography of nature and people-in-process (eyes closed, hands held). Serif typography. Retreat-not-spa aesthetic.",
     "A 50-year-old psychological methodology developed by Bob Hoffman — an intensive process to free people from negative inherited patterns.",
     "The 8-Day Hoffman Process — fully residential, off-grid, no phones, structured group methodology proven across 75,000+ alumni globally.",
     "35-65 adults often in mid-life transition. Lawyers, founders, doctors, parents post-divorce or pre-major-decision.",
     "$4,995-6,500 for 8-day residential. Medical and psychological screening required.",
     "50-year methodology. Certified facilitator training programme. Alumni network as referral engine. No shortcuts (no weekend version offered on principle).",
     "Length signals seriousness. An 8-day commitment filters out tourists. SARA's flagship retreat must demand a real commitment — 7+ days minimum — to attract people who actually transform."),

    # ─── 5. DETOX & RESET ───────────────────────────────────────────────
    ("Detox & Reset",
     "Buchinger Wilhelmi",
     "Therapeutic fasting since 1920.",
     "Clinical-but-warm. White + soft yellow + medical-pale-blue. Family-owned Mediterranean architecture (Marbella) and Bavarian (Überlingen). Staff in lab coats. Fasting-broth imagery.",
     "100+ years and four generations of one family — the world's leading therapeutic fasting clinic with published peer-reviewed research.",
     "The Buchinger Fast — medically supervised juice fast (250-300 calories/day) for 7-21 days, including liver compress, broth rituals, and daily medical check-ins.",
     "European HNWIs 50-70. Often professionals or retired. Returning annually. Many multi-generational guests.",
     "€4,000-12,000 per week (Marbella higher). 7-21 day stays standard.",
     "Four-generation family medical IP. 100-year research base. Peer-reviewed publications. Dual flagship property model.",
     "Family lineage equals trust. A four-generation family business signals consistency the way no chain can. SARA can lean on Montara family ownership as a moat."),

    ("",
     "Vivamayr",
     "Modern Mayr medicine, alpine precision.",
     "White-on-white interiors. Sage + dove grey + glacial blue. Tegernsee and Carinthia lakefront photography. Minimal medical-luxury typography. Chewing-spoon iconography.",
     "The original F.X. Mayr Cure modernised with diagnostic medicine — gut health as the foundation of total health, Austrian lake settings as the reset environment.",
     "Mayr cure — diagnostic intake, individualised diet (chew, chew, chew), manual abdominal treatment, base broth, salt grotto, alkaline-water rituals.",
     "European elite 40-65. Visible Russian/Eastern European, German, and Israeli clientele. Often referred by friends.",
     "€4,500-10,000 per week + diagnostic packages. Minimum 7 nights.",
     "Mayr medicine IP. Lakefront property scarcity. Multi-property strategy (Maria Wörth, Altaussee, Cilli's, London, Marbella). 360+ specialists across portfolio.",
     "Multi-property strategy multiplies brand value. One location is boutique; multiple locations with shared method is a platform. SARA should plan a roadmap to 3-5 properties within 10 years."),

    ("",
     "The Ranch Malibu",
     "Boot-camp wellness for high-output people.",
     "Sun-bleached California ranch. Weathered wood + cream linen + olive green. Hand-lettered farm signage. Golden-hour photography. Equestrian motifs.",
     "A 4 or 7-day fully structured programme — no choice, no menu, no flexibility — designed for high-achieving guests who need someone else to decide for them.",
     "4 AM wake → 4-hour mountain hike → plant-based meals → afternoon strength + yoga + massage → 9 PM lights out. All 100% pre-set.",
     "35-65 Hollywood, tech, and finance executives. Often referred from CAA/WME. Women predominantly but mixed.",
     "$7,000-9,800 per week (4-day Malibu) up to $15,000+ (7-day Hudson Valley). Group size capped (~20).",
     "Fully prescriptive (no choice = no friction). CAA-Hollywood referral network. Multi-location (Malibu + Hudson Valley + Italy). Founder Alex Glasscock as evangelist.",
     "Remove choice for results. High-achievers want to surrender decision-making on holiday. SARA's flagship programme should be a 'no menu' experience — total structural curation."),

    # ─── 6. THE TABLE ───────────────────────────────────────────────────
    ("The Table",
     "Noma",
     "Nordic terroir made into philosophy.",
     "Brutalist + organic. Raw wood, fermented browns, foraged greens. Hand-thrown ceramics. René Redzepi as face. No logo, no neon, no waste imagery.",
     "Cooking what is right here, right now — radical hyper-locality, fermentation as time-travel, and a refusal to import anything that doesn't belong.",
     "Three-season menu (vegetable / seafood / game). 20-course tasting. Foraged ingredients. Fermented kitchen. Noma Projects pantry as extension.",
     "Global gastronomic pilgrims travelling specifically for the meal. Chefs and design-conscious creatives over high-finance crowd.",
     "€725 + €1,200 wine pairing. Seasonal bookings sell out in minutes. Closing 2024 to reopen as 'Noma 3.0' food lab.",
     "René Redzepi singular vision. Nordic terroir narrative invented and owned. Noma Projects products as IP. MAD Symposium as thought-leadership platform.",
     "Terroir as positioning. Noma's whole brand is 'this place, this time'. SARA must claim its terroir — Phuket, Andaman, Thai botanicals — as proprietary, not generic 'Asian wellness'."),

    ("",
     "Flamingo Estate",
     "Garden-to-everything California sensuality.",
     "Tropical pinks, oranges, deep greens. Hand-painted illustrations. Vintage botanical typography. Mediterranean-meets-Hollywood-Hills. Founder Richard Christiansen everywhere.",
     "A working 7-acre estate in Eagle Rock LA whose products (olive oil, soap, candles, honey, tomato sauce) come from the literal land of the founder.",
     "Estate-grown product line — soaps from estate olive oil, honey from estate hives, sauces from estate tomatoes — extended through farmer partnerships.",
     "Design-conscious 30-55 consumers worldwide. Goop-adjacent. Architectural Digest readers. Gifting-heavy purchase behaviour.",
     "$30-150 product range. Estate visits limited and invitation-only. Sold at Goop, Bergdorf, Selfridges.",
     "Real estate (the actual 7-acre property is the brand). Founder personality. Design-led product packaging. Partner-farmer storytelling network.",
     "Sell the land. Flamingo Estate is a property turned into a brand turned into products. SARA's gardens, Pru farm, and Andaman waters should produce branded goods that extend the brand beyond the stay."),

    ("",
     "Single Thread",
     "Kaiseki precision on Sonoma farmland.",
     "Pacific Northwest + Japanese restraint. Charcoal + clay + wood. Custom ceramics by founder Katina Connaughton. Donabe pots. Hand-bound menus. Deep editorial photography.",
     "A 24-acre farm + 11-course Japanese kaiseki menu + 5-room inn — a single thread connecting growing, cooking, eating, and sleeping.",
     "Daily 11-course menu dictated by what the farm produces that morning. Communal counter-style service. Donabe-rice ceremony.",
     "Couples 35-65 from SF/LA/NY treating it as destination weekend. Food-led travellers. Anniversary occasions.",
     "$425 tasting, $700+ with wine. 5 rooms above restaurant $1,200-2,500/night. Reservations 2 months out.",
     "Owned farm (not sourced from). Husband-wife founders Kyle + Katina as personal brand. Integrated property + restaurant + farm. 3 Michelin stars.",
     "Integration is the brand. Single Thread literally is named after its philosophy — farm to table to bed is one continuous chain. SARA's Pru farm + restaurants + villas must read as one philosophy, not three offerings."),

    # ─── 7. BEAUTY & AESTHETICS ─────────────────────────────────────────
    ("Beauty & Aesthetics",
     "Aesop",
     "Apothecary intellectualism in amber glass.",
     "Amber bottle (signature). Brown cream serif (Optima derivative). Institutional-grey + cream packaging. Brutalist-poetic store design (each store architect-led). No model photography. Literary quotations as in-store copy.",
     "A brand that quotes Borges in its store signage — beauty as intellectual practice, products as essays, customer as reader.",
     "In-store hand-wash demonstration — a small ritual of slowness and sensory introduction performed by every staff member.",
     "Design-literate 25-55 worldwide. Architects, writers, gallerists, designers. Gift-buying common.",
     "$25-150 product range. 400+ stores globally, each architecturally distinct. Acquired by L'Oréal 2023 for $2.5B.",
     "Architectural store IP (every store unique). Literary brand voice. Ingredient transparency. 35-year consistency since 1987.",
     "Brand voice is the only true moat. Aesop's product is honestly average — but its voice (serif, sparse, intellectual) is irreproducible. SARA's verbal identity is more valuable than any single product."),

    ("",
     "Augustinus Bader",
     "Stem-cell science from a German laboratory.",
     "Clinical white + cool blue + minimal gold. Doctor-as-spokesperson. Lab photography. Peer-reviewed citations as design elements. Wallpaper-magazine restraint.",
     "Professor Augustinus Bader — biomedical engineer who developed wound-healing technology for burn victims — turns 30 years of stem-cell research into skincare.",
     "The Cream and The Rich Cream — two SKUs only at launch. Applied morning + night. Results in 4 weeks. Minimalist regimen-as-brand.",
     "35-65 women globally. Celebrity adoption (Victoria Beckham, Naomi Campbell) drove early discovery. Clinical-curious skincare buyers.",
     "$290 for The Cream (50ml). $710 for The Rich Cream (50ml). Sold at Net-a-Porter, Bergdorf, brand.com.",
     "Founder credibility (real professor with real research). TFC8® patented technology. Two-product simplicity in a maximalist market. $1B+ valuation in 3 years.",
     "Scientist-founder credibility. Augustinus Bader is the brand because Professor Bader IS real. SARA must surface its scientific advisors — CLP doctors, nutritionists, longevity experts — as the brand's authority figures."),

    ("",
     "Le Labo",
     "Hand-blended perfume with your name on it.",
     "Apothecary lab-coat aesthetic. Mustard yellow + black + raw kraft paper. Hand-typed labels. Smith Corona typewriter. Exposed-piping interiors. No advertising photography.",
     "Every bottle is hand-blended in front of you in the store, labelled with your name and the date — perfume as artisanal craft, not industrial product.",
     "In-store blending ceremony — staff prepare your bottle, write your name on the label, you watch the assembly.",
     "25-50 design-conscious urbanites globally. The 'anti-establishment' perfume customer who rejects Chanel and Dior.",
     "$200-450 per bottle. 'City Exclusives' only sold in their named city (Berlin, Tokyo, etc.). Owned by Estée Lauder since 2014.",
     "City-exclusive geographic scarcity model. Hand-blending ritual. Founder Edouard Roschi & Fabrice Penot artisanal positioning. Kraft-paper consistency since 2006.",
     "Place-bound exclusivity. Le Labo's 'only in Tokyo' City Exclusives are pure scarcity-by-geography. SARA's flagship rituals should be available ONLY in Phuket — never franchised to a sister location."),

    # ─── 8. CULTURE & LEARNING ──────────────────────────────────────────
    ("Culture & Learning",
     "Soho House",
     "Members' clubs for creative industries.",
     "Mid-century-modern + colonial-Victorian + Brooklyn-loft hybrid. Deep greens + burnt orange + vintage prints. Liberty fabric. Gallery-wall art curation. Cookhouse menus.",
     "A club for creative-industry people — founded 1995 by Nick Jones for working in film, fashion, art, and media. No bankers or lawyers admitted on principle (initially).",
     "Membership committee curation — every member is reviewed by industry peers. No amount of money buys admission alone.",
     "28-50 creative professionals globally. 200,000+ members across 40+ houses. Trades on industry homophily.",
     "£2,500-3,500/year + £600-3,200 initiation. 8-year+ waitlists in some cities.",
     "Membership curation IP. 30-year database of creatives. Multi-property network (40+ houses). Hotel-club-restaurant-cinema hybrid model.",
     "Curate the room. The members are the brand. SARA's Member community must be curated for who's in it — not just for revenue. Wrong members destroy the brand."),

    ("",
     "The Battery (San Francisco)",
     "Tech-built private club with cultural soul.",
     "Victorian-meets-tech. Brass + leather + indigo. Restored 1860s building. Dim warm lighting. Custom ceramic plates. Hand-bound member directory. No-photography rule throughout.",
     "Founded by Bebo's Michael Birch — a private club explicitly designed to bring together San Francisco's tech, art, and civic worlds in one room.",
     "Member-curated programming — 200+ events yearly: artist residencies, music performances, dinner salons, civic conversations. 'No work from the tables' rule.",
     "SF/Bay Area founders, artists, philanthropists, journalists. ~3,000 members. Selective committee admission.",
     "$2,800-3,500/year + $1,800 initiation. 'Active member' participation requirement.",
     "Founder's tech-world rolodex. Single-location concentration. Member-curated programming. No-photography policy creating insider mystique.",
     "Programming is the membership. Members don't pay for the rooms — they pay for the calendar. SARA Membership must include a year-round programme of salons, talks, residencies."),

    ("",
     "Annabel's London",
     "Maximalist Mayfair nightclub theatre.",
     "Maximalist florals. Jewel tones. Gold leaf. Hand-painted ceilings. Custom wallpapers (Martin Brudnizki design). No detail unfinished. No surface unadorned.",
     "Founded 1963, reborn 2018. Berkeley Square members' club where dinner, dancing, and theatre belong in one building. 'The most beautiful club in the world.'",
     "Dinner → bar → garden → discotheque arc across five floors and 26,000 sq ft. Each room a different theatrical world.",
     "London ultra-HNW, international visitors, celebrity, royalty. ~9,000 members. Legacy + new tech wealth.",
     "£4,750/year + £4,750 joining fee. Lifetime waitlist.",
     "60+ year history. Mayfair real estate. Martin Brudnizki design IP. Berkeley Square location. Royal-and-celebrity association.",
     "Maximalism done right is unrepeatable. Annabel's didn't go minimal — it went maximal with conviction. SARA can choose richness over restraint as a positioning, provided every detail is executed."),

    # ─── 9. FAMILY & GENERATIONS ────────────────────────────────────────
    ("Family & Generations",
     "Soneva",
     "Barefoot luxury for conscious families.",
     "Sun-bleached driftwood + sea-salt cream + Maldivian-Thai indigo. Hand-drawn maps. Child's-handwriting brand mark. Sandy footprints. 'No News No Shoes' signage.",
     "'Slow life' — Sonu and Eva Shivdasani's philosophy of barefoot, sand-floor luxury that integrates sustainability, children, and wellness as one offering.",
     "'No News, No Shoes' — shoes removed on arrival, returned on departure. Chef's-table kitchen visits. Observatory stargazing. Cinema Paradiso outdoor films.",
     "35-65 affluent families globally. Second-home buyers. Returning multi-generational guests. Sustainability-positive demographic.",
     "$2,500-25,000+/night across Fushi, Jani, Kiri, and Secret in Maldives + Thailand. Branded residences for ownership.",
     "30-year Shivdasani family ownership. Multi-property + branded residences model. Sustainability narrative (waste-to-wealth, no plastic since 2008). Kids' programme integration.",
     "Children are not a problem — they are the proposition. Soneva designed for kids first; adults followed. SARA's multi-generational positioning should be its core, not an afterthought."),

    ("",
     "Soho Farmhouse",
     "Cotswold farm reimagined as members' club.",
     "English countryside + Hamptons hybrid. Whitewashed stone + reclaimed wood + tartan + roaring fires. Cabin architecture. Muddy-boots-friendly luxury.",
     "A 100-acre farm in Oxfordshire converted into a Soho House countryside retreat — town life softened by rural simplicity.",
     "Cabin-to-Main-House daily ritual — wake in your own cabin, walk to Main Barn for breakfast, ride bikes between facilities, cinema barn in the evening.",
     "Soho House members (extension perk). London creative families seeking countryside escape. Multi-generational friendly.",
     "£600-1,500/night for cabins. Soho House membership required (£3,500/year+) plus Farmhouse supplement.",
     "Soho House membership platform (40+ houses) means built-in audience. 100-acre real estate. Cotswold location. Multi-day cabin format.",
     "Cabin format > villa format for repeat stays. Cabins feel personal, ownable, returnable. Multi-gen families want their own building, not a hotel suite. SARA villa programming should follow."),

    ("",
     "Paws Up (Montana)",
     "Glamping ranch for legacy American families.",
     "Big-sky photography. Vintage Americana. Weathered leather + denim + canvas. Taxidermy + cowhide + cast iron. Log-cabin grandeur. Yellowstone-cinema palette.",
     "A 37,000-acre Montana ranch — first to define 'glamping' — where multi-generational American families experience the West with five-star service.",
     "Daily activity menu (rafting, fly-fishing, archery, horseback, ATV, cattle drive, spa, kids' camp) + chuckwagon dinners in the field + private homes (not 'rooms').",
     "Multi-gen American families, often 3-generational. Founders/legacy wealth. Christmas + summer high seasons.",
     "$1,800-4,500/night per person all-inclusive. Private homes $30,000+/night.",
     "37,000-acre owned ranch (irreplaceable scale). 'Glamping' category invention. Multi-gen activity programme. Kids' camp brand within the brand.",
     "'All-inclusive' can be luxury. Paws Up bundles everything (meals, activities, guides, equipment) so guests don't make decisions. SARA's flagship programme should price-bundle similarly for high-decision-fatigue guests."),

    # ─── 10. FARM HOUSE REAL ESTATE ─────────────────────────────────────
    ("Farm House Real Estate",
     "Castello di Reschio",
     "Eight-century estate, family-restored.",
     "Faded ochre + olive green + terracotta + raw linen. Hand-restored architecture. Bolza-family-designed furniture (Reschio Estate brand). Garden topiary. Vintage Fiat 500s. No chrome.",
     "Count Antonio Bolza purchased the 3,700-acre estate in 1994 and spent 30 years restoring 50 farmhouses with his architect son Benedikt. Hotel, homes, and farm are one project.",
     "Farmhouse private rental + truffle hunts + olive harvest + estate-grown ingredients in restaurant + horseback through the estate.",
     "Affluent Anglo-American + European 40-65 families. Design-literate. Often referenced through Architectural Digest, Cabana, House & Garden coverage.",
     "Castello hotel €1,800-7,500/night. Private farmhouse rentals €15,000-50,000/week. ~50 homes available.",
     "3,700-acre family-owned land. 30-year restoration IP. Multi-generational Bolza family as creative directors. Reschio-branded products (wine, olive oil, furniture, perfume).",
     "Build the products. Castello di Reschio sells olive oil, wine, perfume, furniture — brand extensions that travel home with the guest. SARA must build a take-home product line (Pru honey, Tri Vananda balm, SARA tea, etc.)."),

    ("",
     "Babylonstoren",
     "Working Cape Dutch farm, fully brandified.",
     "Cape Dutch whitewashed architecture. Vegetable-garden green. Terracotta. Hand-painted typography. Vintage botanical illustration. Harvest-photography editorial.",
     "A 350-year-old working farm in Franschhoek bought by Karen Roos (former editor of Elle Decoration SA) in 2007 — restored into a farm-hotel-shop-restaurant-spa.",
     "Garden tour daily (the 3.5-acre vegetable garden is the brand's hero asset) + farm restaurant menu dictated by the harvest + spa treatments using estate botanicals.",
     "Affluent international travellers + design-aware Cape Town locals. Cabana / World of Interiors / Travel + Leisure readers.",
     "ZAR 8,500-25,000/night (~$450-1,400). Cottages + farmhouse + Spa Garden Villas. Restaurant + spa + farm shop bookable separately.",
     "Karen Roos editorial-brand instinct. 350-year-old farm provenance. Integrated farm-hotel-shop-restaurant-spa-publishing platform. Expansion to The Newt as proof.",
     "Hire the editor. Babylonstoren works because a magazine editor designed it — every detail merchandised. SARA should bring in editorial brains (Cabana, Wallpaper alumni) to shape its visual world."),

    ("",
     "The Newt in Somerset",
     "English country estate, museum-grade.",
     "Hand-painted apple illustrations. Deep British green + cream. Georgian-house typography. Cyder bottle as design hero. Walled-garden photography. 'No logo on building' restraint.",
     "Karen Roos and Koos Bekker (the Babylonstoren team) acquired Hadspen House in 2013 — 800 acres of Somerset transformed into a cyder-and-garden estate of museum-grade detail.",
     "Apple cyder making (each season's blend a numbered edition) + Roman Villa museum + walled garden + farm-to-restaurant menu + spa + hotel.",
     "London + international design-led travellers. Cyder-curious + garden-curious. Weekend + week-stay mix.",
     "£450-1,200/night across the Farmyard, Hadspen House hotel, and cottages.",
     "Roos/Bekker editorial instinct again. 800-acre estate. Branded cyder production at scale. Roman archaeology asset. Future expansion (Mallorca property coming).",
     "One owner, one vision. The Newt and Babylonstoren both work because of one couple's obsessive direction. SARA cannot be designed by committee — the owner/founder voice must be visible."),
]


# ─── BUILD WORKBOOK ─────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "30-Brand Comparison"

# Styles
thin    = Side(style="thin", color=RULE)
border  = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─── ROW 1: TITLE BAND ──────────────────────────────────────────────────
ws.merge_cells("A1:I1")
title = ws["A1"]
title.value = "SARA — 30-Brand Identity Comparison"
title.font = Font(name=FONT_NAME, size=20, bold=True, color=WHITE)
title.fill = PatternFill("solid", fgColor=FOREST)
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 44

# ─── ROW 2: SUBTITLE ────────────────────────────────────────────────────
ws.merge_cells("A2:I2")
sub = ws["A2"]
sub.value = "World-class brand identity benchmark · 10 categories · 3 brands per category · Strategic intelligence for SARA brand development"
sub.font = Font(name=FONT_NAME, size=11, italic=True, color=INK)
sub.fill = PatternFill("solid", fgColor=SAND)
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 26

# ─── ROW 3: HEADERS ─────────────────────────────────────────────────────
HEADER_ROW = 3
for col_idx, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header)
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=FOREST)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border
ws.row_dimensions[HEADER_ROW].height = 36

# ─── DATA ROWS WITH CATEGORY BANDS ──────────────────────────────────────
row_idx = HEADER_ROW + 1
zebra = False

for entry in DATA:
    category, brand, *cells = entry

    # Insert category band row when category present
    if category:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=9)
        c = ws.cell(row=row_idx, column=1, value=category.upper())
        c.font = Font(name=FONT_NAME, size=11, bold=True, color=GOLD)
        c.fill = PatternFill("solid", fgColor=SAND)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border
        ws.row_dimensions[row_idx].height = 26
        row_idx += 1
        zebra = False

    # Brand row
    fill_color = CREAM if zebra else WHITE
    row_values = [brand] + cells

    for col_idx, value in enumerate(row_values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.border = border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

        if col_idx == 1:  # Brand column
            cell.font = Font(name=FONT_NAME, size=12, bold=True, color=FOREST)
        elif col_idx == 2:  # Essence column — italic
            cell.font = Font(name=FONT_NAME, size=10, italic=True, color=INK)
        elif col_idx == 9:  # Lesson for SARA — gold-accented
            cell.font = Font(name=FONT_NAME, size=10, color=INK)
            cell.fill = PatternFill("solid", fgColor=SAND if not zebra else "F0E9D9")
        else:
            cell.font = Font(name=FONT_NAME, size=10, color=INK)

    ws.row_dimensions[row_idx].height = 165
    row_idx += 1
    zebra = not zebra

# ─── COLUMN WIDTHS ──────────────────────────────────────────────────────
widths = {
    "A": 22,   # Brand
    "B": 28,   # Essence
    "C": 38,   # Visual DNA
    "D": 42,   # Narrative
    "E": 40,   # Ritual
    "F": 32,   # Audience
    "G": 28,   # Price/Access
    "H": 36,   # Moat
    "I": 42,   # Lesson
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# Freeze headers + brand column
ws.freeze_panes = "B4"

# Print setup
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.print_options.horizontalCentered = True
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4
ws.page_margins.top = 0.5
ws.page_margins.bottom = 0.5
ws.print_title_rows = "1:3"

# ─── SECOND SHEET: METHODOLOGY ──────────────────────────────────────────
ws2 = wb.create_sheet("Methodology")
ws2.column_dimensions["A"].width = 26
ws2.column_dimensions["B"].width = 90

methodology = [
    ("SECTION", "CONTENT"),
    ("Purpose", "Strategic identity benchmark for SARA — Montara's new wellness platform brand. Studies 30 world-class brands across 10 categories that SARA's offering ladder touches, with the goal of extracting transferable lessons for SARA's brand development."),
    ("Category logic", "Ten categories reflect the SARA Platform pillars (Vitality, Athletics, Recovery, Mind, Detox, Table, Beauty, Culture, Family) plus one adjacent category (Farm House Real Estate) that informs the SARA villa + Pru integration thesis."),
    ("Brand selection", "Three brands per category, each chosen for distinctive brand identity (not market size). Selection prioritised: (1) ownable visual identity, (2) clear strategic positioning, (3) defensible moat, (4) lesson value for SARA."),
    ("Column definitions", "Essence: 5-word brand definition. Visual DNA: specific design markers. Narrative anchor: the story sold. Signature ritual: proprietary experience. Audience: precise archetype. Price/access: real numbers + gate. Defensibility moat: what's hard to copy. Lesson for SARA: actionable strategic takeaway."),
    ("Use of this document", "Reference for: (1) SARA brand workshop facilitation, (2) creative agency briefing, (3) competitive positioning, (4) trade-off conversations with leadership, (5) onboarding new team members to the SARA aesthetic."),
    ("Next deliverables", "Item 2: Key Insights Summary (8-12 strategic patterns synthesised). Item 3: Research Book (30 brand profiles with 6 sections each)."),
    ("Sources", "Brand websites, IPO filings (Soho House, Aesop/L'Oréal), Architectural Digest, Wallpaper, Monocle, World of Interiors, Cabana, FT How To Spend It, BoF, Robb Report, Travel + Leisure, brand founder interviews (2018-2024)."),
    ("Version", "v1.0 · May 2026 · Tri Vananda / Montara Hospitality Group · Internal Use"),
]

for r, (a, b) in enumerate(methodology, start=1):
    ca = ws2.cell(row=r, column=1, value=a)
    cb = ws2.cell(row=r, column=2, value=b)
    if r == 1:
        ca.font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
        cb.font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
        ca.fill = PatternFill("solid", fgColor=FOREST)
        cb.fill = PatternFill("solid", fgColor=FOREST)
        ca.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        cb.alignment = Alignment(vertical="center", horizontal="left", indent=1)
        ws2.row_dimensions[r].height = 28
    else:
        ca.font = Font(name=FONT_NAME, size=11, bold=True, color=FOREST)
        cb.font = Font(name=FONT_NAME, size=10, color=INK)
        ca.fill = PatternFill("solid", fgColor=SAND)
        ca.alignment = Alignment(vertical="top", horizontal="left", indent=1, wrap_text=True)
        cb.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        ws2.row_dimensions[r].height = 60
    ca.border = border
    cb.border = border

ws2.freeze_panes = "A2"

# Save
out = "/sessions/cool-optimistic-lovelace/mnt/Tri-Vananda/sara-brand-research-comparison-table.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Rows: {row_idx - 1}")
print(f"Brands: 30 across 10 categories")
