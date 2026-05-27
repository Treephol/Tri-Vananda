"""Clean up two SARA-context platform references in xlsx Methodology sheet."""
from openpyxl import load_workbook
from pathlib import Path

path = Path("/sessions/cool-optimistic-lovelace/mnt/Tri-Vananda/sara-brand-research-comparison-table.xlsx")
wb = load_workbook(path)
ws = wb["Methodology"]

# B2: "SARA — Montara's new wellness platform brand" → "SARA — Montara's new wellness brand"
b2 = ws["B2"].value
if b2 and "wellness platform brand" in b2:
    ws["B2"] = b2.replace("wellness platform brand", "wellness brand")
    print(f"B2 updated")

# B3: "SARA Platform pillars" → "SARA Culture pillars"
b3 = ws["B3"].value
if b3 and "SARA Platform pillars" in b3:
    ws["B3"] = b3.replace("SARA Platform pillars", "SARA Culture pillars")
    print(f"B3 updated")

wb.save(path)
print("Saved.")
